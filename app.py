"""
AGM Motor Bakım Merkezi — Streamlit uygulaması.
Tek dosyalık, çok kullanıcılı, rol tabanlı, fotoğraf destekli motor bakım
takip sistemi. MongoDB Atlas'a bağlanır. Streamlit Community Cloud'da
Docker/Render gerekmeden doğrudan çalışır.
"""
import base64
import hashlib
import io
import json
import os
from datetime import datetime, date, timedelta

import pandas as pd
import pymongo
import streamlit as st
from PIL import Image

st.set_page_config(page_title="AGM Motor Bakım Merkezi", page_icon="🔧", layout="wide")

KRITIK_ESIK = 100
YAKLASIYOR_ESIK = 250
STATUS_LABELS = {"gecikmis": "🔴 Gecikmiş", "kritik": "🟠 Kritik", "yaklasiyor": "🟡 Yaklaşıyor", "normal": "🟢 Normal"}
STATUS_COLORS = {"gecikmis": "#ef4a52", "kritik": "#f2994a", "yaklasiyor": "#f0c93d", "normal": "#2fb374"}
ROLES = ["yonetici", "planlamaci", "teknisyen", "goruntuleyici"]
ROLE_LABELS = {"yonetici": "Yönetici", "planlamaci": "Planlamacı", "teknisyen": "Teknisyen", "goruntuleyici": "Görüntüleyici"}

# ============================================================
# Görsel tema — grafit zemin, amber/teal vurgu, endüstriyel tipografi
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Barlow+Condensed:wght@600;700&family=JetBrains+Mono:wght@500;600;700&display=swap');

:root{
  --bg:#0f1319; --panel:#171d25; --panel2:#1f2730; --border:#2a323c; --border-lt:#374252;
  --text:#eef1f5; --muted:#8b96a3; --faint:#5b6572;
  --amber:#e8952f; --amber-soft:rgba(232,149,47,.14);
  --teal:#3fb5c4; --teal-soft:rgba(63,181,196,.14);
  --red:#ef4a52; --orange:#f2994a; --yellow:#f0c93d; --green:#33c98a;
}

html, body, [class*="css"]{ font-family:'Inter',sans-serif; }
.stApp{ background:radial-gradient(circle at 20% 0%, #141a22 0%, #0f1319 55%); color:var(--text); }

/* Streamlit varsayılan üst şerit / footer'ı sadeleştir */
#MainMenu{ visibility:hidden; }
footer{ visibility:hidden; }
header[data-testid="stHeader"]{ background:transparent; }

/* Başlıklar */
h1, h2, h3{ font-family:'Barlow Condensed',sans-serif !important; text-transform:uppercase; letter-spacing:.4px; color:var(--text) !important; }
h3{ font-size:19px !important; font-weight:700 !important; border-bottom:1px solid var(--border); padding-bottom:8px; margin-bottom:14px !important;}

/* Sidebar */
section[data-testid="stSidebar"]{ background:#12161d; border-right:1px solid var(--border); }
section[data-testid="stSidebar"] .stRadio label{
  padding:9px 12px; border-radius:10px; margin-bottom:2px; font-weight:600; font-size:13.5px;
}
section[data-testid="stSidebar"] .stRadio [data-baseweb="radio"]{ display:none; }
section[data-testid="stSidebar"] .stRadio label:has(input:checked){
  background:var(--amber-soft); color:var(--amber) !important; border:1px solid rgba(232,149,47,.35);
}

/* Kartlar / container'lar */
div[data-testid="stVerticalBlockBorderWrapper"]{
  background:var(--panel); border:1px solid var(--border) !important; border-radius:14px;
}
div[data-testid="stExpander"]{ background:var(--panel); border:1px solid var(--border); border-radius:14px; overflow:hidden;}
div[data-testid="stExpander"] summary{ font-weight:700; font-size:13.5px; }

/* Metric */
div[data-testid="stMetric"]{
  background:var(--panel); border:1px solid var(--border); border-radius:14px; padding:12px 14px;
}
div[data-testid="stMetricValue"]{ font-family:'JetBrains Mono',monospace !important; color:var(--amber) !important; }
div[data-testid="stMetricLabel"]{ color:var(--muted) !important; font-size:11.5px !important; text-transform:uppercase; letter-spacing:.4px; }

/* Butonlar */
.stButton button{
  border-radius:10px !important; font-weight:700 !important; border:1px solid var(--border-lt) !important;
  background:var(--panel2) !important; color:var(--text) !important;
}
.stButton button[kind="primary"]{
  background:linear-gradient(180deg,#f0a23f,#e8952f) !important; color:#1a1206 !important; border:none !important;
  box-shadow:0 6px 18px rgba(232,149,47,.28);
}
.stDownloadButton button{ border-radius:10px !important; font-weight:700 !important; }

/* Input alanları */
.stTextInput input, .stTextArea textarea, .stNumberInput input, .stDateInput input{
  background:var(--panel2) !important; color:var(--text) !important; border:1px solid var(--border) !important; border-radius:10px !important;
}
div[data-baseweb="select"] > div{ background:var(--panel2) !important; border-color:var(--border) !important; border-radius:10px !important; }

/* Sekmeler */
.stTabs [data-baseweb="tab-list"]{ gap:4px; background:#12161d; padding:4px; border-radius:10px; border:1px solid var(--border);}
.stTabs [data-baseweb="tab"]{ border-radius:8px; font-weight:700; font-size:12.5px; color:var(--faint); }
.stTabs [aria-selected="true"]{ background:var(--amber) !important; color:#161006 !important; }

/* Dataframe / tablo */
div[data-testid="stDataFrame"]{ border-radius:12px; overflow:hidden; border:1px solid var(--border); }

/* JetBrains Mono ile sayısal hislendirme için yardımcı sınıf */
.mono-num{ font-family:'JetBrains Mono',monospace; font-weight:700; }

/* Durum rozeti (özel HTML kartlarda kullanılır) */
.status-pill{ display:inline-flex; align-items:center; gap:5px; padding:3px 10px; border-radius:999px; font-size:11px; font-weight:800; }
.status-pill.gecikmis{ background:rgba(239,74,82,.14); color:var(--red); }
.status-pill.kritik{ background:rgba(242,153,74,.14); color:var(--orange); }
.status-pill.yaklasiyor{ background:rgba(240,201,61,.14); color:var(--yellow); }
.status-pill.normal{ background:rgba(51,201,138,.14); color:var(--green); }
</style>
""", unsafe_allow_html=True)


# ============================================================
# Veritabanı bağlantısı
# ============================================================
@st.cache_resource
def get_db():
    uri = st.secrets["MONGO_URI"]
    db_name = st.secrets.get("MONGO_DB_NAME", "agm_bakim")
    client = pymongo.MongoClient(uri, serverSelectionTimeoutMS=8000)
    return client[db_name]


db = get_db()
engines_col = db["engines"]
types_col = db["maintenance_types"]
records_col = db["maintenance_records"]
users_col = db["users"]
oil_analyses_col = db["oil_analyses"]
pressure_readings_col = db["pressure_readings"]
equipment_info_col = db["equipment_info"]


def seed_if_empty():
    """
    Veritabanı boşsa (veya önceki bir deneme yarıda kalmışsa) V10 dosyasından
    çıkarılan gerçek verilerle doldurur. 'upsert' kullanır — aynı anda iki kez
    tetiklense veya kısmen daha önce çalışmış olsa bile hata vermez.
    """
    seed_path = os.path.join(os.path.dirname(__file__), "seed_data.json")
    with open(seed_path, encoding="utf-8") as f:
        data = json.load(f)

    if engines_col.count_documents({}) < len(data["engines"]):
        now = datetime.utcnow()
        for name, info in data["engines"].items():
            engines_col.update_one(
                {"_id": name},
                {"$setOnInsert": {
                    "name": name, "hours": info["hours"], "load_kw": info.get("load", 0),
                    "updated_at": now, "history": [{"date": now.isoformat(), "hours": info["hours"]}],
                }},
                upsert=True,
            )

    expected_type_count = 1 + len(data["maintTypes"])  # +1 = yağ
    if types_col.count_documents({}) < expected_type_count:
        oil_states = {name: {"last_maintenance_hour": rec["changeHour"], "period_hours": rec["maxHours"]}
                      for name, rec in data["oil"].items()}
        types_col.update_one(
            {"_id": "oil"},
            {"$setOnInsert": {"key": "oil", "label": "Yağ Değişimi", "default_period_hours": 700, "engine_states": oil_states}},
            upsert=True,
        )
        for mt in data["maintTypes"]:
            states = {name: {"last_maintenance_hour": rec["lastHour"], "period_hours": rec["period"]}
                      for name, rec in mt["perEngine"].items()}
            default_period = next(iter(states.values()))["period_hours"] if states else 0
            types_col.update_one(
                {"_id": mt["key"]},
                {"$setOnInsert": {"key": mt["key"], "label": mt["label"],
                                   "default_period_hours": default_period, "engine_states": states}},
                upsert=True,
            )


def seed_equipment_info():
    """Motorlar_hava_filtre_ve_kaver_tipleri dosyasındaki referans bilgileri, koleksiyon boşsa yükler."""
    if equipment_info_col.count_documents({}) > 0:
        return
    path = os.path.join(os.path.dirname(__file__), "equipment_info.json")
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    for name, info in data.items():
        equipment_info_col.update_one({"_id": name}, {"$setOnInsert": {**info, "engine_name": name}}, upsert=True)


def seed_pressure_history():
    """KARTER_FARK_BASINÇLARI dosyasındaki geçmiş okumaları, koleksiyon boşsa yükler."""
    if pressure_readings_col.count_documents({}) > 0:
        return
    path = os.path.join(os.path.dirname(__file__), "karter_history.json")
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        readings = json.load(f)
    docs = []
    for r in readings:
        docs.append({
            "engine_id": r["engine"], "engine_name": r["engine"],
            "reading_date": datetime.fromisoformat(r["date"]),
            "load_kw": r["load"], "pressure_bar": r["pressure"], "status": r.get("status"),
            "new_type": r.get("new_type", False), "note": None,
            "uploaded_by": "V10 içe aktarma", "created_at": datetime.utcnow(),
        })
    if docs:
        pressure_readings_col.insert_many(docs)


seed_if_empty()
seed_equipment_info()
seed_pressure_history()


# ============================================================
# Durum hesaplama
# ============================================================
def remaining_hours(engine_hours, last_hour, period):
    return period - (engine_hours - last_hour)


def status_for(remaining):
    if remaining <= 0:
        return "gecikmis"
    if remaining <= KRITIK_ESIK:
        return "kritik"
    if remaining <= YAKLASIYOR_ESIK:
        return "yaklasiyor"
    return "normal"


@st.cache_data(ttl=15)
def build_items():
    engines = {e["_id"]: e for e in engines_col.find()}
    types = list(types_col.find())
    items = []
    for t in types:
        states = t.get("engine_states", {})
        applicable = list(states.keys()) if states else list(engines.keys())
        for eng_id in applicable:
            engine = engines.get(eng_id)
            if not engine:
                continue
            state = states.get(eng_id, {})
            last_hour = state.get("last_maintenance_hour", 0)
            period = state.get("period_hours", t["default_period_hours"])
            remaining = remaining_hours(engine["hours"], last_hour, period)
            items.append({
                "engine_id": eng_id, "engine_name": engine["name"], "type_key": t["key"], "type_label": t["label"],
                "engine_hours": engine["hours"], "last_hour": last_hour, "period": period,
                "remaining": remaining, "status": status_for(remaining),
            })
    return items, engines, types


def engine_sort_key(name):
    digits = "".join(ch for ch in name if ch.isdigit())
    return int(digits) if digits else 0


def style_status_df(df):
    """DataFrame'de 'Durum' sütunu varsa, hücreleri gecikmiş/kritik/yaklaşıyor/normal
    renklerine göre boyayan bir Styler döndürür. Yoksa DataFrame'i olduğu gibi bırakır."""
    if "Durum" not in df.columns:
        return df
    color_map = {label: STATUS_COLORS[key] for key, label in STATUS_LABELS.items()}

    def _color(val):
        c = color_map.get(val)
        return f"color:{c}; font-weight:700;" if c else ""

    styler = df.style
    map_fn = getattr(styler, "map", None) or styler.applymap
    return map_fn(_color, subset=["Durum"])


def _gauge_ring_svg(remaining, period, color, size=44):
    import math
    r = size / 2 - 4
    circumference = 2 * math.pi * r
    pct = max(0.0, min(1.0, remaining / period)) if period else 0.0
    dash = circumference * pct
    return (
        f'<svg width="{size}" height="{size}" style="flex-shrink:0;">'
        f'<circle cx="{size/2}" cy="{size/2}" r="{r}" fill="none" stroke="#2a323c" stroke-width="4"/>'
        f'<circle cx="{size/2}" cy="{size/2}" r="{r}" fill="none" stroke="{color}" stroke-width="4" '
        f'stroke-dasharray="{circumference:.1f}" stroke-dashoffset="{circumference-dash:.1f}" '
        f'stroke-linecap="round" transform="rotate(-90 {size/2} {size/2})"/></svg>'
    )


def _engine_badge_html(name, size=32):
    num = engine_sort_key(name)
    fs = round(size * 0.36)
    return (
        f'<div style="width:{size}px;height:{size}px;border-radius:9px;display:flex;align-items:center;'
        f'justify-content:center;background:linear-gradient(155deg,#232d3a,#171d25);border:1px solid var(--border);'
        f'font-family:\'JetBrains Mono\',monospace;font-weight:700;font-size:{fs}px;color:var(--amber);flex-shrink:0;">{num}</div>'
    )


def render_gauge_cards(rows):
    """rows: title, subtitle, status, remaining, period, value_label, unit_label, badge_name (opsiyonel) içeren sözlük listesi.
    Her satırı gösterge halkalı, durum renkli bir kart olarak çizer — mouse ile kaydırmaya gerek kalmadan tümü tek seferde görünür."""
    if not rows:
        st.info("Kayıt bulunamadı.")
        return
    html = '<div style="display:flex;flex-direction:column;gap:7px;">'
    for r in rows:
        color = STATUS_COLORS[r["status"]]
        ring = _gauge_ring_svg(r["remaining"], r["period"], color)
        badge = _engine_badge_html(r["badge_name"]) if r.get("badge_name") else ""
        pill = f'<span class="status-pill {r["status"]}" style="margin-top:5px;">{STATUS_LABELS[r["status"]]}</span>'
        html += f'''
        <div style="display:flex;align-items:center;gap:12px;background:var(--panel);border:1px solid var(--border);
                    border-radius:14px;padding:11px 12px;">
          {badge}{ring}
          <div style="flex:1;min-width:0;">
            <div style="font-size:13px;font-weight:700;color:var(--text);">{r['title']}</div>
            <div style="font-size:11px;color:var(--faint);margin-top:2px;">{r['subtitle']}</div>
            {pill}
          </div>
          <div style="text-align:right;flex-shrink:0;">
            <div class="mono-num" style="font-size:14px;color:{color};">{r['value_label']}</div>
            <div style="font-size:9px;color:var(--faint);letter-spacing:.3px;">{r['unit_label']}</div>
          </div>
        </div>'''
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


def render_load_cards(engines_list):
    """Motor yükü kartlarını yatay kaydırmalı bir şerit olarak çizer."""
    html = '<div style="display:flex;gap:8px;overflow-x:auto;padding-bottom:10px;">'
    for e in engines_list:
        html += f'''
        <div style="display:flex;flex-direction:column;align-items:center;gap:5px;flex-shrink:0;width:74px;
                    padding:10px 6px;border-radius:12px;background:var(--panel);border:1px solid var(--border);">
          {_engine_badge_html(e["name"], size=28)}
          <span style="font-size:10px;color:var(--muted);font-weight:600;">{e["name"]}</span>
          <span class="mono-num" style="font-size:13px;color:var(--teal);">{e.get("load_kw",0):,.0f}</span>
          <span style="font-size:8px;color:var(--faint);letter-spacing:.3px;">kW</span>
        </div>'''
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


def estimate_daily_usage(engine_doc):
    """Motorun geçmiş çalışma saati kayıtlarından günlük ortalama kullanım (saat/gün) hesaplar."""
    if not engine_doc:
        return None
    history = engine_doc.get("history", [])
    if len(history) < 2:
        return None
    try:
        history_sorted = sorted(history, key=lambda h: h["date"])
        first, last = history_sorted[0], history_sorted[-1]
        first_dt = datetime.fromisoformat(first["date"])
        last_dt = datetime.fromisoformat(last["date"])
        span_days = (last_dt - first_dt).total_seconds() / 86400
        if span_days < 0.5:
            return None
        delta_hours = last["hours"] - first["hours"]
        if delta_hours <= 0:
            return None
        return delta_hours / span_days
    except Exception:
        return None


def delete_button(collection, doc_id, key_suffix, current_user, owner_id=None, on_delete=None):
    """Silme yetkisi: yönetici/planlamacı her kaydı, diğer roller yalnızca
    kendi oluşturdukları kaydı silebilir. İki adımlı onay ister."""
    can_delete = current_user["role"] in ("yonetici", "planlamaci") or (owner_id == current_user["_id"])
    if not can_delete:
        return
    confirm_key = f"confirm_del_{key_suffix}"
    if st.session_state.get(confirm_key):
        c1, c2 = st.columns(2)
        if c1.button("❌ Vazgeç", key=f"cancel_{key_suffix}", use_container_width=True):
            st.session_state[confirm_key] = False
            st.rerun()
        if c2.button("🗑️ Evet, Sil", key=f"confirmed_{key_suffix}", type="primary", use_container_width=True):
            collection.delete_one({"_id": doc_id})
            if on_delete:
                on_delete()
            st.cache_data.clear()
            st.success("Kayıt silindi.")
            st.session_state[confirm_key] = False
            st.rerun()
    else:
        if st.button("🗑️ Sil", key=f"del_{key_suffix}"):
            st.session_state[confirm_key] = True
            st.rerun()


def recompute_last_maintenance(engine_id, type_key):
    """Bir motor + bakım türü için 'son bakım saati'ni, o türe ait tüm kayıtlar
    arasındaki en güncel (en yüksek) çalışma saatinden yeniden hesaplar.
    Kayıt eklendiğinde, düzenlendiğinde veya silindiğinde çağrılır — böylece
    kalan saat hesabı her zaman doğru kalır."""
    remaining = list(records_col.find({"engine_id": engine_id, "type_key": type_key}))
    if not remaining:
        return
    max_hour = max(rec["hour_at_completion"] for rec in remaining)
    types_col.update_one(
        {"_id": type_key},
        {"$set": {f"engine_states.{engine_id}.last_maintenance_hour": max_hour}},
        upsert=True,
    )


def update_engine_hours(engine_id, new_hours):
    """Motorun merkezi (güncel) çalışma saatini günceller ve geçmişe bir kayıt ekler."""
    stamp = datetime.utcnow()
    engines_col.update_one(
        {"_id": engine_id},
        {"$set": {"hours": new_hours, "updated_at": stamp},
         "$push": {"history": {"date": stamp.isoformat(), "hours": new_hours}}},
    )


# ============================================================
# Kimlik doğrulama (basit, Streamlit oturumu üzerinden)
# ============================================================
def hash_password(password, salt=None):
    if salt is None:
        salt = os.urandom(16).hex()
    h = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), 100_000).hex()
    return h, salt


def verify_password(password, salt, expected_hash):
    h, _ = hash_password(password, salt)
    return h == expected_hash


def create_user(full_name, email, password, role):
    salt_hash, salt = hash_password(password)
    users_col.insert_one({
        "_id": email.lower().strip(), "full_name": full_name, "email": email.lower().strip(),
        "password_hash": salt_hash, "salt": salt, "role": role, "active": True,
        "created_at": datetime.utcnow(),
    })


def login_view():
    st.markdown("## 🔧 AGM Motor Bakım Merkezi")
    user_count = users_col.count_documents({})

    if user_count == 0:
        st.info("Sistemde henüz kullanıcı yok. İlk kaydolan kişi otomatik olarak **yönetici** olur.")
        with st.form("ilk_kayit"):
            full_name = st.text_input("Adınız Soyadınız")
            email = st.text_input("E-posta")
            password = st.text_input("Şifre", type="password")
            submitted = st.form_submit_button("Yönetici Hesabı Oluştur", use_container_width=True)
            if submitted:
                if not full_name or not email or len(password) < 6:
                    st.error("Lütfen tüm alanları doldurun (şifre en az 6 karakter olmalı).")
                else:
                    create_user(full_name, email, password, "yonetici")
                    st.success("Hesap oluşturuldu! Şimdi giriş yapabilirsiniz.")
                    st.rerun()
        return

    tab_login, tab_register = st.tabs(["Giriş Yap", "Yeni Hesap (Teknisyen)"])

    with tab_login:
        with st.form("giris"):
            email = st.text_input("E-posta")
            password = st.text_input("Şifre", type="password")
            submitted = st.form_submit_button("Giriş Yap", use_container_width=True)
            if submitted:
                user = users_col.find_one({"_id": email.lower().strip()})
                if not user or not verify_password(password, user["salt"], user["password_hash"]):
                    st.error("E-posta veya şifre hatalı.")
                elif not user.get("active", True):
                    st.error("Hesabınız devre dışı bırakılmış. Yöneticinizle iletişime geçin.")
                else:
                    st.session_state.user = user
                    st.rerun()

    with tab_register:
        st.caption("Saha teknisyenleri buradan kendi hesabını oluşturabilir. Yönetici veya planlamacı yetkisi gerekiyorsa, giriş yaptıktan sonra bir yöneticiden 'Kullanıcılar' sayfasından rolünüzü değiştirmesini isteyin.")
        with st.form("kayit"):
            full_name = st.text_input("Adınız Soyadınız", key="reg_name")
            email = st.text_input("E-posta", key="reg_email")
            password = st.text_input("Şifre", type="password", key="reg_pass")
            submitted = st.form_submit_button("Hesap Oluştur", use_container_width=True)
            if submitted:
                if not full_name or not email or len(password) < 6:
                    st.error("Lütfen tüm alanları doldurun (şifre en az 6 karakter olmalı).")
                elif users_col.find_one({"_id": email.lower().strip()}):
                    st.error("Bu e-posta zaten kayıtlı.")
                else:
                    create_user(full_name, email, password, "teknisyen")
                    st.success("Hesap oluşturuldu! 'Giriş Yap' sekmesinden giriş yapabilirsiniz.")


# ============================================================
# Sayfalar
# ============================================================
def render_stat_cards(counts):
    order = [("gecikmis", "Gecikmiş"), ("kritik", "Kritik"), ("yaklasiyor", "Yaklaşıyor"), ("normal", "Normal")]
    cards = ""
    for key, label in order:
        color = STATUS_COLORS[key]
        cards += f"""
        <div style="background:var(--panel); border:1px solid var(--border); border-radius:14px;
                    padding:13px 14px; position:relative; overflow:hidden; flex:1; min-width:110px;">
          <div style="position:absolute; top:0; left:0; width:3px; height:100%; background:{color};"></div>
          <div style="font-size:10.5px; font-weight:700; letter-spacing:.5px; color:var(--muted); text-transform:uppercase;">{label}</div>
          <div style="font-family:'JetBrains Mono',monospace; font-size:26px; font-weight:700; margin-top:4px; color:{color};">{counts[key]}</div>
        </div>"""
    st.markdown(f'<div style="display:flex; gap:8px; flex-wrap:wrap; margin-bottom:6px;">{cards}</div>', unsafe_allow_html=True)


def page_dashboard():
    items, engines, types = build_items()

    counts = {"gecikmis": 0, "kritik": 0, "yaklasiyor": 0, "normal": 0}
    for i in items:
        counts[i["status"]] += 1
    render_stat_cards(counts)

    st.markdown("### Bakım Bildirimleri")
    for status_key, header in [("gecikmis", "🔴 Gecikmiş"), ("kritik", "🟠 Kritik"), ("yaklasiyor", "🟡 Yaklaşıyor")]:
        subset = [i for i in items if i["status"] == status_key]
        with st.expander(f"{header} ({len(subset)})", expanded=(status_key == "gecikmis" and len(subset) > 0)):
            if not subset:
                st.caption("Kayıt yok.")
                continue
            by_type = {}
            for i in subset:
                by_type.setdefault(i["type_label"], []).append(i)
            rows = sorted(by_type.items(), key=lambda kv: -len(kv[1]))
            breakdown_df = pd.DataFrame([{"Bakım Türü": label, "Motor Sayısı": len(lst)} for label, lst in rows])
            st.dataframe(breakdown_df, use_container_width=True, hide_index=True,
                         height=min(35 * (len(breakdown_df) + 1) + 3, 500))

            type_labels = [label for label, _ in rows]
            detail_choice = st.selectbox("Hangi bakım türünün motorlarını görmek istersiniz?",
                                          ["Seçiniz..."] + type_labels, key=f"detail_{status_key}")
            if detail_choice != "Seçiniz...":
                detail_items = sorted(by_type[detail_choice], key=lambda i: i["remaining"])
                unit = "SAAT GECİKME" if status_key == "gecikmis" else "SAAT KALDI"
                card_rows = [{
                    "title": i["engine_name"],
                    "subtitle": f"Motor saati {i['engine_hours']:,.0f} sa · Son bakım {i['last_hour']:,.0f} sa · Çalışılan {i['engine_hours']-i['last_hour']:,.0f} sa",
                    "status": i["status"], "remaining": i["remaining"], "period": i["period"],
                    "value_label": ("+" if i["remaining"] <= 0 else "") + f"{abs(i['remaining']):,.0f}",
                    "unit_label": unit, "badge_name": i["engine_name"],
                } for i in detail_items]
                render_gauge_cards(card_rows)

    st.markdown("### Motor Yükleri")
    load_rows = sorted(engines.values(), key=lambda e: engine_sort_key(e["name"]))
    total_load = sum(e.get("load_kw", 0) for e in load_rows)
    avg_load = total_load / len(load_rows) if load_rows else 0
    lc1, lc2 = st.columns(2)
    lc1.metric("Toplam Yük", f"{total_load:,.0f} kW")
    lc2.metric("Ortalama Yük", f"{avg_load:,.0f} kW")
    render_load_cards(load_rows)

    st.markdown("### Bakım Türüne Göre Görüntüle")
    type_options = ["Tümü"] + sorted({i["type_label"] for i in items})
    type_choice = st.selectbox("Bakım türü seç", type_options, key="dash_type_filter")
    status_choice = st.selectbox("Durum (opsiyonel)", ["Tümü", "Gecikmiş", "Kritik", "Yaklaşıyor", "Normal"], key="dash_status_filter")
    filter_map = {"Gecikmiş": "gecikmis", "Kritik": "kritik", "Yaklaşıyor": "yaklasiyor", "Normal": "normal"}

    rows = items
    if type_choice != "Tümü":
        rows = [i for i in rows if i["type_label"] == type_choice]
    if status_choice != "Tümü":
        rows = [i for i in rows if i["status"] == filter_map[status_choice]]
    rows = sorted(rows, key=lambda i: i["remaining"])

    card_rows = [{
        "title": r["engine_name"],
        "subtitle": f"{r['type_label']} · {r['engine_hours']:,.0f} sa motor saati · Çalışılan {r['engine_hours']-r['last_hour']:,.0f} sa",
        "status": r["status"], "remaining": r["remaining"], "period": r["period"],
        "value_label": ("+" if r["remaining"] <= 0 else "") + f"{abs(r['remaining']):,.0f}",
        "unit_label": "SAAT GECİKME" if r["remaining"] <= 0 else "SAAT KALDI",
        "badge_name": r["engine_name"],
    } for r in rows]
    render_gauge_cards(card_rows)


def page_hours_update():
    st.markdown("### Motor Çalışma Saatlerini / Yüklerini Güncelle")
    st.caption("Bu ekrandan güncellediğiniz saatler, tüm bakım türlerindeki kalan süreleri otomatik olarak yeniden hesaplar.")

    engines = sorted(engines_col.find(), key=lambda e: engine_sort_key(e["name"]))
    with st.form("saat_guncelle"):
        new_hours = {}
        new_loads = {}
        for e in engines:
            c1, c2 = st.columns(2)
            new_hours[e["_id"]] = c1.number_input(f"{e['name']} — Saat", value=float(e["hours"]), step=1.0, key=f"hr_{e['_id']}")
            new_loads[e["_id"]] = c2.number_input(f"{e['name']} — Yük (kW)", value=float(e.get("load_kw", 0)), step=10.0, key=f"ld_{e['_id']}")
        submitted = st.form_submit_button("💾 Kaydet", use_container_width=True, type="primary")

    if submitted:
        stamp = datetime.utcnow()
        changed = 0
        for e in engines:
            new_h = new_hours[e["_id"]]
            new_l = new_loads[e["_id"]]
            set_fields = {}
            push_history = False
            if new_h != e["hours"]:
                set_fields["hours"] = new_h
                push_history = True
            if new_l != e.get("load_kw", 0):
                set_fields["load_kw"] = new_l
            if set_fields:
                set_fields["updated_at"] = stamp
                update_op = {"$set": set_fields}
                if push_history:
                    update_op["$push"] = {"history": {"date": stamp.isoformat(), "hours": new_h}}
                engines_col.update_one({"_id": e["_id"]}, update_op)
                changed += 1
        st.cache_data.clear()
        st.success(f"{changed} motor güncellendi." if changed else "Değişiklik yapılmadı.")


def page_engines():
    items, engines, types = build_items()
    st.markdown("### Motorlar")
    st.caption("Bir motoru açarak o motora ait tüm bakım türlerini ve durumlarını görebilirsiniz.")

    query = st.text_input("Motor ara", placeholder="örn. AGM 12")
    sort_by = st.radio("Sırala", ["Durum", "Motor No", "Çalışma Saati", "Yük"], horizontal=True)
    status_order = {"gecikmis": 0, "kritik": 1, "yaklasiyor": 2, "normal": 3}
    status_icon = {"gecikmis": "🔴", "kritik": "🟠", "yaklasiyor": "🟡", "normal": "🟢"}

    rows = []
    for name, e in engines.items():
        if query and query.lower() not in name.lower():
            continue
        eng_items = sorted([i for i in items if i["engine_id"] == name], key=lambda i: i["remaining"])
        worst = eng_items[0] if eng_items else None
        status = worst["status"] if worst else "normal"
        rows.append({"name": name, "hours": e["hours"], "load": e.get("load_kw", 0),
                      "status": status, "items": eng_items,
                      "worst_remaining": worst["remaining"] if worst else 999999})

    if sort_by == "Durum":
        rows.sort(key=lambda r: (status_order[r["status"]], r["worst_remaining"]))
    elif sort_by == "Motor No":
        rows.sort(key=lambda r: engine_sort_key(r["name"]))
    elif sort_by == "Çalışma Saati":
        rows.sort(key=lambda r: -r["hours"])
    else:
        rows.sort(key=lambda r: -r["load"])

    if not rows:
        st.info("Eşleşen motor bulunamadı.")
        return

    for r in rows:
        label = f"{status_icon[r['status']]} {r['name']} · {r['hours']:,.0f} sa · {r['load']:,.0f} kW"
        with st.expander(label):
            if not r["items"]:
                st.caption("Bu motor için tanımlı bakım türü yok.")
                continue
            card_rows = [{
                "title": i["type_label"],
                "subtitle": f"Periyot {i['period']:,.0f} sa · Son bakım {i['last_hour']:,.0f} sa · Çalışılan {i['engine_hours']-i['last_hour']:,.0f} sa",
                "status": i["status"], "remaining": i["remaining"], "period": i["period"],
                "value_label": ("+" if i["remaining"] <= 0 else "") + f"{abs(i['remaining']):,.0f}",
                "unit_label": "SAAT GECİKME" if i["remaining"] <= 0 else "SAAT KALDI",
            } for i in r["items"]]
            render_gauge_cards(card_rows)


def page_types():
    items, engines, types = build_items()
    st.markdown("### Bakım Türleri")

    type_options = {t["label"]: t["key"] for t in sorted(types, key=lambda t: t["label"])}
    selected_label = st.selectbox("Bakım türü seç", list(type_options.keys()))
    selected_key = type_options[selected_label]
    status_filter = st.selectbox("Durum filtresi", ["Tümü", "Gecikmiş", "Kritik", "Yaklaşıyor", "Normal"], key="types_status")
    filter_map = {"Gecikmiş": "gecikmis", "Kritik": "kritik", "Yaklaşıyor": "yaklasiyor", "Normal": "normal"}

    rows = [i for i in items if i["type_key"] == selected_key]
    if status_filter != "Tümü":
        rows = [r for r in rows if r["status"] == filter_map[status_filter]]
    rows.sort(key=lambda r: r["remaining"])

    card_rows = [{
        "title": r["engine_name"],
        "subtitle": f"Motor saati {r['engine_hours']:,.0f} sa · Son bakım {r['last_hour']:,.0f} sa · Periyot {r['period']:,.0f} sa · Çalışılan {r['engine_hours']-r['last_hour']:,.0f} sa",
        "status": r["status"], "remaining": r["remaining"], "period": r["period"],
        "value_label": ("+" if r["remaining"] <= 0 else "") + f"{abs(r['remaining']):,.0f}",
        "unit_label": "SAAT GECİKME" if r["remaining"] <= 0 else "SAAT KALDI",
        "badge_name": r["engine_name"],
    } for r in rows]
    render_gauge_cards(card_rows)


def compress_photo(uploaded_file, max_dim=720, quality=65):
    img = Image.open(uploaded_file)
    img = img.convert("RGB")
    w, h = img.size
    if max(w, h) > max_dim:
        if w > h:
            new_w, new_h = max_dim, int(h * max_dim / w)
        else:
            new_w, new_h = int(w * max_dim / h), max_dim
        img = img.resize((new_w, new_h))
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


MAX_VIDEO_MB = 15
VIDEO_MIME_MAP = {"mp4": "video/mp4", "mov": "video/quicktime", "webm": "video/webm",
                   "m4v": "video/mp4", "avi": "video/x-msvideo"}


def encode_video(uploaded_file):
    """Videoyu base64'e çevirir. Video sıkıştırma yapılmaz (ek kütüphane
    gerektirir) — bu yüzden boyut sınırı MongoDB'nin belge başına 16MB
    limitine göre belirlenmiştir. Sınırı aşan dosyalar reddedilir."""
    if uploaded_file.size > MAX_VIDEO_MB * 1024 * 1024:
        return None, f"'{uploaded_file.name}' dosyası {MAX_VIDEO_MB}MB sınırını aşıyor, eklenmedi. Daha kısa bir klip deneyin."
    ext = uploaded_file.name.rsplit(".", 1)[-1].lower() if "." in uploaded_file.name else "mp4"
    mime = VIDEO_MIME_MAP.get(ext, "video/mp4")
    data_b64 = base64.b64encode(uploaded_file.read()).decode("utf-8")
    return {"data_b64": data_b64, "filename": uploaded_file.name, "mime": mime}, None


def page_complete_maintenance(current_user):
    if current_user["role"] == "goruntuleyici":
        st.warning("Görüntüleyici rolü bakım tamamlayamaz.")
        return

    items, engines, types = build_items()
    st.markdown("### Bakım Tamamla")

    engine_names = sorted(engines.keys(), key=engine_sort_key)
    engine_name = st.selectbox("Motor", engine_names)

    # Bu motor için zaten tanımlı olan bakım türleri
    eng_items = sorted([i for i in items if i["engine_id"] == engine_name], key=lambda i: i["remaining"])
    applicable_keys = {i["type_key"] for i in eng_items}

    # Excel'deki TÜM bakım türlerini göster — motor için henüz tanımlı olmayanlar
    # da seçilebilir (seçildiğinde bu motor için yeni bir bakım takibi başlatılır).
    all_types_sorted = sorted(types, key=lambda t: t["label"])
    label_to_type = {}
    for t in all_types_sorted:
        if t["key"] in applicable_keys:
            item = next(i for i in eng_items if i["type_key"] == t["key"])
            label = f"{t['label']} · {STATUS_LABELS[item['status']]} · {round(item['remaining'])} sa"
        else:
            label = f"{t['label']} · ⚪ Bu motor için tanımlı değil"
        label_to_type[label] = t

    chosen_label = st.selectbox("Bakım türü", list(label_to_type.keys()))
    chosen_type = label_to_type[chosen_label]
    chosen_key = chosen_type["key"]
    is_new_for_engine = chosen_key not in applicable_keys

    engine_hours_now = engines[engine_name]["hours"]

    if is_new_for_engine:
        st.warning(f"**{chosen_type['label']}**, {engine_name} için Excel'de tanımlı değildi. Bu kaydı eklersen bu motor için yeni bir bakım takibi başlatılır.")
        period = st.number_input("Periyodik bakım saati (bu motor için)", min_value=1.0,
                                  value=float(chosen_type["default_period_hours"]), step=100.0)
        last_hour_before = 0.0
    else:
        chosen_item = next(i for i in eng_items if i["type_key"] == chosen_key)
        period = chosen_item["period"]
        last_hour_before = chosen_item["last_hour"]
        st.info(f"**{chosen_type['label']}** — Motor saati: {engine_hours_now} · Son bakım: {last_hour_before} · Periyot: {period}")

    st.markdown("---")
    backdated = st.checkbox("📅 Geçmişe dönük kayıt (bu bakım geçmişte yapıldı, bugün değil)")
    if backdated:
        record_date = st.date_input("Bakımın yapıldığı tarih", value=date.today(), max_value=date.today())
    else:
        record_date = date.today()

    record_hours = st.number_input(
        "O anki motor çalışma saati", min_value=0.0, value=float(engine_hours_now), step=1.0,
        help="Bakımın yapıldığı andaki motor çalışma saatini girin. Bu değer motorun güncel çalışma saatinden BÜYÜKSE motorun güncel saati de bu değere güncellenir; küçük veya eşitse yalnızca bu bakım kaydına yazılır."
    )

    # Karter fark basıncı — krankcase filtresi ve intercooler bakımlarında ölçülür
    pressure_reading = None
    if chosen_key in ("krank", "intercooler"):
        pressure_reading = st.number_input("Fark Basıncı (bar)", min_value=0.0, step=0.1, format="%.2f")

    note = st.text_area("Ölçüm / Teknik Açıklama (opsiyonel)", help="Ölçüm değerleri, gözlemler, teknik detaylar")
    tech_note = st.text_area("Bakımcı Notu (opsiyonel)", help="Bakımı yapan kişinin genel notu/yorumu")
    photos = st.file_uploader("Fotoğraf ekle (opsiyonel, birden fazla seçebilirsiniz)",
                               type=["jpg", "jpeg", "png", "webp"], accept_multiple_files=True)
    camera_photo = st.camera_input("Ya da doğrudan fotoğraf çek (opsiyonel)")
    videos = st.file_uploader(f"Video ekle (opsiyonel, kısa klipler — her biri en fazla {MAX_VIDEO_MB}MB)",
                               type=["mp4", "mov", "webm", "m4v"], accept_multiple_files=True)

    st.markdown("---")
    other_items = [i for i in eng_items if i["type_key"] != chosen_key]
    extra_selected = []
    if other_items:
        st.caption("Bazen bir bakımı yaparken diğerlerini de yapmış oluyorsunuz (örn. siloksan temizliğiyle birlikte intercooler/yağ). Bu bakımla birlikte tamamlanan başka bakımlar varsa işaretleyin — hepsi aynı saat ve tarihle kaydedilir.")
        extra_labels = {f"{i['type_label']} · {STATUS_LABELS[i['status']]} · {round(i['remaining'])} sa": i for i in other_items}
        picked = st.multiselect("Bu bakımla birlikte tamamlanan diğer bakımlar (opsiyonel)", list(extra_labels.keys()))
        extra_selected = [extra_labels[p] for p in picked]

    if st.button("✅ Bakımı Tamamla", use_container_width=True, type="primary"):
        photos_b64 = []
        all_photos = list(photos) if photos else []
        if camera_photo is not None:
            all_photos.append(camera_photo)
        for p in all_photos:
            try:
                photos_b64.append(compress_photo(p))
            except Exception:
                st.warning("Bir fotoğraf işlenemedi, atlandı.")

        videos_data = []
        for v in (videos or []):
            encoded, err = encode_video(v)
            if err:
                st.warning(err)
            elif encoded:
                videos_data.append(encoded)

        record_datetime = datetime.combine(record_date, datetime.now().time())
        record = {
            "engine_id": engine_name, "engine_name": engine_name,
            "type_key": chosen_key, "type_label": chosen_type["label"],
            "hour_at_completion": record_hours, "note": note, "technician_note": tech_note,
            "photos_b64": photos_b64, "videos": videos_data,
            "technician_id": current_user["_id"], "technician_name": current_user["full_name"],
            "created_at": record_datetime, "backdated": backdated,
        }
        if pressure_reading is not None:
            record["pressure_reading"] = pressure_reading
        records_col.insert_one(record)

        # Bu bakım türünün periyodunu (yeni tanımlanan tür ise veya değiştiyse) kaydet,
        # ardından 'son bakım saati'ni bu motor + tür için var olan TÜM kayıtlardan
        # yeniden hesapla — böylece kalan saat her zaman doğru kalır.
        types_col.update_one(
            {"_id": chosen_key},
            {"$set": {f"engine_states.{engine_name}.period_hours": period}},
            upsert=True,
        )
        recompute_last_maintenance(engine_name, chosen_key)

        # Aynı anda tamamlandığı işaretlenen diğer bakımlar için de aynı saat/tarih/
        # teknisyen ile ayrı kayıtlar oluştur.
        completed_labels = [chosen_type["label"]]
        for extra in extra_selected:
            extra_record = {
                "engine_id": engine_name, "engine_name": engine_name,
                "type_key": extra["type_key"], "type_label": extra["type_label"],
                "hour_at_completion": record_hours, "note": note, "technician_note": tech_note,
                "photos_b64": photos_b64, "videos": videos_data,
                "technician_id": current_user["_id"], "technician_name": current_user["full_name"],
                "created_at": record_datetime, "backdated": backdated,
                "grouped_with": chosen_type["label"],
            }
            records_col.insert_one(extra_record)
            recompute_last_maintenance(engine_name, extra["type_key"])
            completed_labels.append(extra["type_label"])

        # Girilen saat, motorun GÜNCEL çalışma saatinden büyükse motorun güncel
        # saatini de bu değere günceller (daha yeni bir bilgi). Küçük veya eşitse
        # motorun güncel saatine dokunmaz, yalnızca bu bakım kaydına yazılır —
        # böylece geçmişe dönük ya da eski bir saat girişi güncel saati geri
        # düşürmez.
        if record_hours > engine_hours_now:
            update_engine_hours(engine_name, record_hours)

        st.cache_data.clear()
        st.success(f"{', '.join(completed_labels)} bakımı {engine_name} için kaydedildi.")
        st.rerun()


def page_records(current_user):
    st.markdown("### Bakım Kayıtları")

    engine_names = sorted({e["name"] for e in engines_col.find()}, key=engine_sort_key)
    type_labels = sorted({t["label"] for t in types_col.find()})

    c1, c2 = st.columns(2)
    engine_filter = c1.selectbox("Motora göre filtrele", ["Tümü"] + engine_names, key="rec_engine_filter")
    type_filter = c2.selectbox("Bakım türüne göre filtrele", ["Tümü"] + type_labels, key="rec_type_filter")

    query = {}
    if engine_filter != "Tümü":
        query["engine_id"] = engine_filter
    if type_filter != "Tümü":
        query["type_label"] = type_filter

    records = list(records_col.find(query).sort("created_at", -1).limit(500))
    st.caption(f"{len(records)} kayıt")
    if not records:
        st.info("Kayıt bulunamadı.")
        return

    for r in records:
        with st.container(border=True):
            photos = r.get("photos_b64") or ([r["photo_b64"]] if r.get("photo_b64") else [])
            if photos:
                cols = st.columns([1, 3])
                with cols[0]:
                    if len(photos) == 1:
                        st.image(base64.b64decode(photos[0]), use_container_width=True)
                    else:
                        photo_cols = st.columns(min(len(photos), 3))
                        for idx, p in enumerate(photos):
                            with photo_cols[idx % len(photo_cols)]:
                                st.image(base64.b64decode(p), use_container_width=True)
                info_col = cols[1]
            else:
                info_col = st.container()
            with info_col:
                title = f"**{r['type_label']}** · {r['engine_name']}"
                if r.get("backdated"):
                    title += " · 📅 geçmişe dönük"
                st.markdown(title)
                st.caption(f"{r['created_at'].strftime('%d.%m.%Y')} · {r['hour_at_completion']} sa okumasında · {r.get('technician_name','')}")
                if r.get("pressure_reading") is not None:
                    st.caption(f"📈 Fark Basıncı: {r['pressure_reading']} bar")
                if r.get("note"):
                    st.caption(f"📝 {r['note']}")
                if r.get("technician_note"):
                    st.caption(f"🗒️ Bakımcı Notu: {r['technician_note']}")

                for v in (r.get("videos") or []):
                    st.video(base64.b64decode(v["data_b64"]), format=v.get("mime", "video/mp4"))

                can_edit = current_user["role"] in ("yonetici", "planlamaci") or current_user["_id"] == r.get("technician_id")
                bc1, bc2 = st.columns(2)
                if can_edit:
                    with bc1:
                        if st.button("✏️ Düzenle", key=f"editbtn_{r['_id']}"):
                            st.session_state[f"edit_{r['_id']}"] = not st.session_state.get(f"edit_{r['_id']}", False)
                            st.rerun()
                with bc2:
                    delete_button(records_col, r["_id"], f"rec_{r['_id']}", current_user, owner_id=r.get("technician_id"),
                                  on_delete=lambda eng=r["engine_id"], tk=r["type_key"]: recompute_last_maintenance(eng, tk))

                if can_edit and st.session_state.get(f"edit_{r['_id']}"):
                    edit_maintenance_record(r, photos)


def edit_maintenance_record(r, current_photos):
    st.markdown("---")
    st.caption("Düzenleme modu")
    new_hours = st.number_input("Motor Çalışma Saati (bu bakımın yapıldığı andaki)",
                                 min_value=0.0, value=float(r["hour_at_completion"]), step=1.0,
                                 key=f"ed_hours_{r['_id']}",
                                 help="Bu bakım kaydına ait çalışma saatini düzeltir. Geçmişe dönük olmayan kayıtlarda motorun güncel çalışma saatini de günceller.")
    new_note = st.text_area("Ölçüm / Teknik Açıklama", value=r.get("note", ""), key=f"ed_note_{r['_id']}")
    new_tech_note = st.text_area("Bakımcı Notu", value=r.get("technician_note", ""), key=f"ed_tnote_{r['_id']}")
    new_pressure = None
    if r.get("type_key") in ("krank", "intercooler") or r.get("pressure_reading") is not None:
        new_pressure = st.number_input("Fark Basıncı (bar)", value=float(r.get("pressure_reading") or 0.0),
                                        step=0.1, format="%.2f", key=f"ed_press_{r['_id']}")

    keep_photos = list(current_photos)
    if current_photos:
        st.caption("Mevcut fotoğraflar — kaldırmak istediklerinizi işaretleyin")
        remove_flags = []
        cols = st.columns(min(len(current_photos), 4))
        for idx, p in enumerate(current_photos):
            with cols[idx % len(cols)]:
                st.image(base64.b64decode(p), use_container_width=True)
                remove_flags.append(st.checkbox("Kaldır", key=f"ed_rm_{r['_id']}_{idx}"))
        keep_photos = [p for p, rm in zip(current_photos, remove_flags) if not rm]

    new_photos = st.file_uploader("Yeni fotoğraf ekle (opsiyonel, birden fazla seçebilirsiniz)",
                                   type=["jpg", "jpeg", "png", "webp"], accept_multiple_files=True,
                                   key=f"ed_upload_{r['_id']}")
    new_camera = st.camera_input("Ya da doğrudan fotoğraf çek", key=f"ed_camera_{r['_id']}")

    current_videos = r.get("videos") or []
    keep_videos = list(current_videos)
    if current_videos:
        st.caption("Mevcut videolar — kaldırmak istediklerinizi işaretleyin")
        keep_videos = []
        for idx, v in enumerate(current_videos):
            st.video(base64.b64decode(v["data_b64"]), format=v.get("mime", "video/mp4"))
            remove = st.checkbox(f"Bu videoyu kaldır ({v.get('filename','video')})", key=f"ed_vrm_{r['_id']}_{idx}")
            if not remove:
                keep_videos.append(v)

    new_videos = st.file_uploader(f"Yeni video ekle (opsiyonel, her biri en fazla {MAX_VIDEO_MB}MB)",
                                   type=["mp4", "mov", "webm", "m4v"], accept_multiple_files=True,
                                   key=f"ed_video_upload_{r['_id']}")

    c1, c2 = st.columns(2)
    if c1.button("Vazgeç", key=f"ed_cancel_{r['_id']}", use_container_width=True):
        st.session_state[f"edit_{r['_id']}"] = False
        st.rerun()
    if c2.button("💾 Kaydet", key=f"ed_save_{r['_id']}", type="primary", use_container_width=True):
        added_photos = []
        all_new = list(new_photos) if new_photos else []
        if new_camera is not None:
            all_new.append(new_camera)
        for p in all_new:
            try:
                added_photos.append(compress_photo(p))
            except Exception:
                st.warning("Bir fotoğraf işlenemedi, atlandı.")

        added_videos = []
        for v in (new_videos or []):
            encoded, err = encode_video(v)
            if err:
                st.warning(err)
            elif encoded:
                added_videos.append(encoded)

        update = {
            "hour_at_completion": new_hours, "note": new_note,
            "technician_note": new_tech_note, "photos_b64": keep_photos + added_photos,
            "videos": keep_videos + added_videos,
        }
        if new_pressure is not None:
            update["pressure_reading"] = new_pressure
        records_col.update_one({"_id": r["_id"]}, {"$set": update})

        # Saat değiştiyse, bu motor + bakım türü için 'son bakım saati'ni
        # tüm kayıtlardan yeniden hesapla.
        if new_hours != r["hour_at_completion"]:
            recompute_last_maintenance(r["engine_id"], r["type_key"])
            # Yeni saat, motorun GÜNCEL çalışma saatinden büyükse motorun güncel
            # saatini de günceller. Küçük veya eşitse dokunmaz.
            current_engine = engines_col.find_one({"_id": r["engine_id"]})
            if current_engine and new_hours > current_engine["hours"]:
                update_engine_hours(r["engine_id"], new_hours)

        st.session_state[f"edit_{r['_id']}"] = False
        st.cache_data.clear()
        st.success("Kayıt güncellendi.")
        st.rerun()


def page_maintenance_forecast():
    items, engines, types = build_items()
    st.markdown("### Bakım Tarihi Tahmini")
    st.caption("Motorun geçmiş çalışma saati kayıtlarından günlük ortalama kullanım hesaplanır; bu hıza göre bakımın hangi takvim tarihinde geleceği tahmin edilir. En az iki farklı tarihte saat kaydı olan motorlar için tahmin yapılabilir.")

    type_options = ["Tümü"] + sorted({i["type_label"] for i in items})
    type_choice = st.selectbox("Bakım türü seç", type_options, key="forecast_type")

    rows = items if type_choice == "Tümü" else [i for i in items if i["type_label"] == type_choice]
    if not rows:
        st.info("Kayıt bulunamadı.")
        return

    forecast_rows = []
    for r in rows:
        engine_doc = engines.get(r["engine_id"])
        daily = estimate_daily_usage(engine_doc)
        if daily and daily > 0:
            days_left = r["remaining"] / daily
            est_date = date.today() + timedelta(days=days_left)
            est_date_str = est_date.strftime("%d.%m.%Y")
            sort_key = days_left
            daily_str = f"{daily:,.1f} sa/gün"
        else:
            est_date_str = "Tahmin edilemiyor"
            sort_key = 10**9
            daily_str = "yetersiz veri"
        forecast_rows.append({
            "title": r["engine_name"],
            "subtitle": f"{r['type_label']} · Kalan {r['remaining']:,.0f} sa · {daily_str}",
            "status": r["status"], "remaining": r["remaining"], "period": r["period"],
            "value_label": est_date_str, "unit_label": "TAHMİNİ TARİH",
            "badge_name": r["engine_name"], "_sort": sort_key,
        })
    forecast_rows.sort(key=lambda r: r["_sort"])
    render_gauge_cards(forecast_rows)


def page_hours_history():
    st.markdown("### Motor Saat Geçmişi")
    engines = sorted(engines_col.find(), key=lambda e: engine_sort_key(e["name"]))
    names = [e["name"] for e in engines]
    selected = st.selectbox("Motor seç", names)
    engine = next(e for e in engines if e["name"] == selected)
    history = sorted(engine.get("history", []), key=lambda h: h["date"])

    if len(history) < 2:
        st.info("Bu motor için henüz yeterli geçmiş kaydı yok.")
        return

    df = pd.DataFrame(history)
    df["date"] = pd.to_datetime(df["date"])
    df = df.rename(columns={"date": "Tarih", "hours": "Saat"}).set_index("Tarih")
    st.line_chart(df)

    total_delta = history[-1]["hours"] - history[0]["hours"]
    span_days = max(1, (pd.to_datetime(history[-1]["date"]) - pd.to_datetime(history[0]["date"])).days)
    c1, c2, c3 = st.columns(3)
    c1.metric("Toplam Artış", f"{total_delta:,.0f} sa")
    c2.metric("Günlük Ortalama", f"{total_delta/span_days:,.1f} sa")
    c3.metric("Kayıt Sayısı", len(history))

    table_df = pd.DataFrame(history)[::-1]
    table_df["date"] = pd.to_datetime(table_df["date"]).dt.strftime("%d.%m.%Y")
    st.dataframe(table_df.rename(columns={"date": "Tarih", "hours": "Saat"}), use_container_width=True, hide_index=True)


def page_intervals():
    st.markdown("### Bakım Aralıkları")
    st.caption("Her motor + bakım türü için art arda tamamlanan bakımlar arasında geçen saat farkını gösterir.")
    records = list(records_col.find().sort("created_at", 1))
    if not records:
        st.info("Henüz tamamlanmış bakım yok. İlk bakımı kaydettiğinizde burada birikmeye başlayacak.")
        return

    groups = {}
    for r in records:
        key = (r["engine_name"], r["type_key"])
        groups.setdefault(key, {"label": r["type_label"], "entries": []})
        groups[key]["entries"].append(r)

    engine_filter = st.selectbox("Motora göre filtrele", ["Tümü"] + sorted({r["engine_name"] for r in records}, key=engine_sort_key))

    for (engine_name, type_key), g in sorted(groups.items(), key=lambda kv: engine_sort_key(kv[0][0])):
        if engine_filter != "Tümü" and engine_name != engine_filter:
            continue
        entries = g["entries"]
        st.markdown(f"**{engine_name} — {g['label']}**")
        rows = []
        prev = None
        for i, e in enumerate(entries):
            delta = None if prev is None else e["hour_at_completion"] - prev["hour_at_completion"]
            rows.append({
                "Tarih": e["created_at"].strftime("%d.%m.%Y"), "Motor Saati": e["hour_at_completion"],
                "Teknisyen": e.get("technician_name", ""),
                "Aralık": "MİLAD" if delta is None else f"{delta:,.0f} sa",
            })
            prev = e
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        if len(entries) >= 2:
            avg = (entries[-1]["hour_at_completion"] - entries[0]["hour_at_completion"]) / (len(entries) - 1)
            st.caption(f"Ortalama aralık: **{avg:,.0f} saat**")
        st.divider()


def page_export():
    st.markdown("### Excel Dışa / İçe Aktar")

    st.markdown("#### 📤 Rapor İndir")
    if st.button("Excel raporu oluştur"):
        items, engines, types = build_items()
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            eng_df = pd.DataFrame([{"MOTOR": e["name"], "MOTOR ÇALIŞMA SAATİ": e["hours"], "YÜK (kW)": e.get("load_kw", 0)}
                                    for e in sorted(engines.values(), key=lambda e: engine_sort_key(e["name"]))])
            eng_df.to_excel(writer, sheet_name="Motor Saatleri", index=False)

            summary_rows = sorted(items, key=lambda i: i["remaining"])
            summary_df = pd.DataFrame([{
                "MOTOR": i["engine_name"], "BAKIM TÜRÜ": i["type_label"], "MOTOR SAATİ": i["engine_hours"],
                "SON BAKIM SAATİ": i["last_hour"], "PERİYOT": i["period"],
                "KALAN SAAT": round(i["remaining"], 1), "DURUM": STATUS_LABELS[i["status"]],
            } for i in summary_rows])
            summary_df.to_excel(writer, sheet_name="Bakım Özeti", index=False)

            by_type = {}
            for i in items:
                by_type.setdefault(i["type_label"], []).append(i)
            for label, rows in by_type.items():
                rows = sorted(rows, key=lambda r: engine_sort_key(r["engine_name"]))
                df = pd.DataFrame([{
                    "MOTOR": r["engine_name"], "MOTOR SAATİ": r["engine_hours"], "SON BAKIM SAATİ": r["last_hour"],
                    "PERİYOT": r["period"], "KALAN SAAT": round(r["remaining"], 1), "DURUM": STATUS_LABELS[r["status"]],
                } for r in rows])
                df.to_excel(writer, sheet_name=label[:31], index=False)

        st.download_button("İndir (.xlsx)", buf.getvalue(),
                            file_name=f"AGM_Motor_Bakim_Raporu_{date.today().isoformat()}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("#### 📥 Motor Saatlerini / Yüklerini İçe Aktar")
    st.caption("'MOTOR' ve 'MOTOR ÇALIŞMA SAATİ' sütunlarını içeren bir Excel dosyası yükleyin. Dosyada ayrıca 'YÜK' (kW) sütunu varsa, motor yükleri de aynı anda güncellenir.")
    up = st.file_uploader("Excel dosyası seç", type=["xlsx", "xls"], key="import_file")
    if up is not None and st.button("İçe aktar"):
        df = pd.read_excel(up)
        cols = {c.upper(): c for c in df.columns}
        name_col = next((v for k, v in cols.items() if "MOTOR" in k and "SAAT" not in k and "YÜK" not in k), None)
        hour_col = next((v for k, v in cols.items() if "SAAT" in k), None)
        load_col = next((v for k, v in cols.items() if "YÜK" in k), None)
        if not name_col or not hour_col:
            st.error("MOTOR ve MOTOR ÇALIŞMA SAATİ sütunları bulunamadı.")
        else:
            stamp = datetime.utcnow()
            updated = 0
            for _, row in df.iterrows():
                name = str(row[name_col]).strip()
                hours = row[hour_col]
                load_val = row[load_col] if load_col else None
                existing = engines_col.find_one({"_id": name})
                if not existing:
                    continue
                set_fields = {"updated_at": stamp}
                push_history = False
                if not pd.isna(hours) and float(hours) != existing["hours"]:
                    set_fields["hours"] = float(hours)
                    push_history = True
                if load_col is not None and not pd.isna(load_val):
                    set_fields["load_kw"] = float(load_val)
                if len(set_fields) > 1:
                    update_op = {"$set": set_fields}
                    if push_history:
                        update_op["$push"] = {"history": {"date": stamp.isoformat(), "hours": float(hours)}}
                    engines_col.update_one({"_id": name}, update_op)
                    updated += 1
            st.cache_data.clear()
            st.success(f"{updated} motor için çalışma saati güncellendi.")


def page_oil_analyses(current_user):
    st.markdown("### Yağ Analizleri (Laboratuvar Raporları)")
    st.caption("Laboratuvarda yapılan yağ analizi PDF raporlarını motor bazında saklayın.")

    engines = sorted(engines_col.find(), key=lambda e: engine_sort_key(e["name"]))
    engine_names = [e["name"] for e in engines]

    if current_user["role"] != "goruntuleyici":
        with st.expander("➕ Yeni analiz raporu ekle", expanded=False):
            engine_name = st.selectbox("Motor", engine_names, key="oil_pdf_engine")
            analysis_date = st.date_input("Numune / analiz tarihi", value=date.today(), max_value=date.today())
            result = st.selectbox("Genel değerlendirme", ["İyi", "Dikkat", "Kötü"], key="oil_pdf_result")
            note = st.text_area("Not (opsiyonel)", key="oil_pdf_note")
            pdf_file = st.file_uploader("PDF raporu", type=["pdf"], key="oil_pdf_file")

            if st.button("Raporu Kaydet", type="primary"):
                if pdf_file is None:
                    st.error("Lütfen bir PDF dosyası seçin.")
                elif pdf_file.size > 10 * 1024 * 1024:
                    st.error("Dosya 10MB sınırını aşıyor. Daha küçük bir dosya deneyin.")
                else:
                    pdf_b64 = base64.b64encode(pdf_file.read()).decode("utf-8")
                    oil_analyses_col.insert_one({
                        "engine_id": engine_name, "engine_name": engine_name,
                        "analysis_date": datetime.combine(analysis_date, datetime.min.time()),
                        "result": result, "note": note, "pdf_b64": pdf_b64, "pdf_filename": pdf_file.name,
                        "uploaded_by": current_user["full_name"], "uploaded_by_id": current_user["_id"],
                        "created_at": datetime.utcnow(),
                    })
                    st.success(f"{engine_name} için analiz raporu kaydedildi.")
                    st.rerun()

    st.markdown("---")
    filter_engine = st.selectbox("Motora göre filtrele", ["Tümü"] + engine_names, key="oil_pdf_filter")
    query = {} if filter_engine == "Tümü" else {"engine_id": filter_engine}
    analyses = list(oil_analyses_col.find(query).sort("analysis_date", -1))

    if not analyses:
        st.info("Henüz analiz raporu eklenmemiş.")
        return

    result_icons = {"İyi": "🟢", "Dikkat": "🟡", "Kötü": "🔴"}
    for a in analyses:
        with st.container(border=True):
            c1, c2 = st.columns([4, 1])
            with c1:
                st.markdown(f"**{a['engine_name']}** · {result_icons.get(a.get('result',''), '')} {a.get('result','')}")
                st.caption(f"Analiz tarihi: {a['analysis_date'].strftime('%d.%m.%Y')} · Yükleyen: {a.get('uploaded_by','')}")
                if a.get("note"):
                    st.caption(f"📝 {a['note']}")
                delete_button(oil_analyses_col, a["_id"], f"oil_{a['_id']}", current_user, owner_id=a.get("uploaded_by_id"))
            with c2:
                if a.get("pdf_b64"):
                    st.download_button("📄 PDF İndir", base64.b64decode(a["pdf_b64"]),
                                        file_name=a.get("pdf_filename", "analiz.pdf"),
                                        mime="application/pdf", key=f"dl_{a['_id']}")


def page_pressure_readings(current_user):
    st.markdown("### Karter Fark Basıncı")
    st.caption("Motorların düzenli karter fark basıncı ve yük ölçümlerini kaydedin ve zaman içindeki değişimi izleyin.")

    engines = sorted(engines_col.find(), key=lambda e: engine_sort_key(e["name"]))
    engine_names = [e["name"] for e in engines]

    tab_new, tab_history, tab_import = st.tabs(["➕ Yeni Ölçüm", "📈 Geçmiş / Grafik", "📥 Excel'den İçe Aktar"])

    with tab_new:
        if current_user["role"] == "goruntuleyici":
            st.warning("Görüntüleyici rolü ölçüm ekleyemez.")
        else:
            reading_date = st.date_input("Ölçüm tarihi", value=date.today(), max_value=date.today(), key="pr_date")
            st.caption("Her motor için yük (kW) ve fark basıncını (bar) girin. Bakımda olan motorlar için kutucuğu işaretleyin.")
            entries = {}
            for e in engines:
                with st.container(border=True):
                    c1, c2, c3, c4 = st.columns([2, 2, 2, 1.5])
                    c1.markdown(f"**{e['name']}**")
                    under_maint = c4.checkbox("Bakımda/Yedek", key=f"pr_maint_{e['_id']}")
                    load_val = c2.number_input("Yük (kW)", key=f"pr_load_{e['_id']}", step=10.0, disabled=under_maint, label_visibility="collapsed", placeholder="Yük (kW)")
                    pressure_val = c3.number_input("Fark Basıncı (bar)", key=f"pr_press_{e['_id']}", step=0.1, format="%.2f", disabled=under_maint, label_visibility="collapsed", placeholder="Fark Basıncı")
                    entries[e["name"]] = (load_val, pressure_val, under_maint)

            if st.button("💾 Tüm Ölçümleri Kaydet", type="primary", use_container_width=True):
                stamp = datetime.combine(reading_date, datetime.now().time())
                docs = []
                for name, (load_val, pressure_val, under_maint) in entries.items():
                    if under_maint:
                        docs.append({"engine_id": name, "engine_name": name, "reading_date": stamp,
                                      "load_kw": None, "pressure_bar": None, "status": "BAKIMDA",
                                      "new_type": False, "note": None,
                                      "uploaded_by": current_user["full_name"], "uploaded_by_id": current_user["_id"],
                                      "created_at": datetime.utcnow()})
                    elif load_val or pressure_val:
                        docs.append({"engine_id": name, "engine_name": name, "reading_date": stamp,
                                      "load_kw": load_val or None, "pressure_bar": pressure_val or None, "status": None,
                                      "new_type": False, "note": None,
                                      "uploaded_by": current_user["full_name"], "uploaded_by_id": current_user["_id"],
                                      "created_at": datetime.utcnow()})
                if docs:
                    pressure_readings_col.insert_many(docs)
                    st.success(f"{len(docs)} motor için ölçüm kaydedildi.")
                    st.rerun()
                else:
                    st.warning("Kaydedilecek bir değer girilmedi.")

    with tab_history:
        selected = st.selectbox("Motor seç", engine_names, key="pr_hist_engine")
        readings = list(pressure_readings_col.find({"engine_id": selected}).sort("reading_date", 1))
        numeric_readings = [r for r in readings if r.get("pressure_bar") is not None]

        if len(numeric_readings) >= 2:
            df = pd.DataFrame([{"Tarih": r["reading_date"], "Fark Basıncı (bar)": r["pressure_bar"]} for r in numeric_readings])
            df = df.set_index("Tarih")
            st.line_chart(df)

        if readings:
            table_rows = []
            for r in reversed(readings):
                table_rows.append({
                    "Tarih": r["reading_date"].strftime("%d.%m.%Y"),
                    "Yük (kW)": r.get("load_kw") if r.get("load_kw") is not None else "-",
                    "Fark Basıncı (bar)": r.get("pressure_bar") if r.get("pressure_bar") is not None else "-",
                    "Durum": r.get("status") or "-",
                })
            st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

            with st.expander("Tek tek kayıt sil"):
                for r in reversed(readings):
                    c1, c2 = st.columns([4, 1])
                    c1.caption(f"{r['reading_date'].strftime('%d.%m.%Y')} · Basınç: {r.get('pressure_bar','-')} · Yük: {r.get('load_kw','-')}")
                    with c2:
                        delete_button(pressure_readings_col, r["_id"], f"pr_{r['_id']}", current_user, owner_id=r.get("uploaded_by_id"))
        else:
            st.info("Bu motor için henüz ölçüm kaydı yok.")

    with tab_import:
        st.caption("KARTER_FARK_BASINÇLARI.xlsx ile aynı yapıdaki bir dosyayı yükleyerek geçmiş ölçümleri toplu ekleyebilirsiniz. Her sayfa adı bir tarih (GG.AA.YYYY) olmalıdır.")
        up = st.file_uploader("Excel dosyası seç", type=["xlsx"], key="pr_import_file")
        if up is not None and st.button("İçe Aktar", key="pr_import_btn"):
            try:
                added = import_pressure_excel(up, current_user)
                st.success(f"{added} ölçüm kaydı eklendi.")
                st.rerun()
            except Exception as e:
                st.error(f"Dosya okunamadı: {e}")


def import_pressure_excel(uploaded_file, current_user):
    from openpyxl import load_workbook
    wb = load_workbook(uploaded_file, data_only=True)

    def to_number(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def norm(name):
        s = str(name).strip().replace("-", " ")
        return " ".join(s.split())

    docs = []
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        try:
            d, m, y = sheetname.split(".")
            sheet_date = datetime(int(y), int(m), int(d))
        except Exception:
            continue

        header_row = None
        for r in range(1, 6):
            for c in range(1, 12):
                if ws.cell(r, c).value == "MOTOR NO":
                    header_row = r
                    break
            if header_row:
                break
        if header_row is None:
            continue

        header_cells = [(c, str(ws.cell(header_row, c).value).strip())
                         for c in range(1, 15) if ws.cell(header_row, c).value is not None]
        blocks, current = [], None
        for c, label in header_cells:
            if label == "MOTOR NO":
                if current:
                    blocks.append(current)
                current = {"motor_col": c, "load_col": None, "pressure_col": None, "end_col": c}
            elif current:
                if label == "YÜK":
                    current["load_col"] = c
                elif label == "KARTER FARK BASINCI":
                    current["pressure_col"] = c
                current["end_col"] = c
        if current:
            blocks.append(current)

        for r in range(header_row + 1, ws.max_row + 1):
            for b in blocks:
                engine_raw = ws.cell(r, b["motor_col"]).value
                if not engine_raw or "AGM" not in str(engine_raw).upper():
                    continue
                engine = norm(engine_raw)
                if not engines_col.find_one({"_id": engine}):
                    continue
                load_val = ws.cell(r, b["load_col"]).value if b["load_col"] else None
                pressure_val = ws.cell(r, b["pressure_col"]).value if b["pressure_col"] else None
                status = None
                for v in (load_val, pressure_val):
                    if v is not None and to_number(v) is None:
                        status = str(v).strip().upper()
                docs.append({
                    "engine_id": engine, "engine_name": engine, "reading_date": sheet_date,
                    "load_kw": to_number(load_val), "pressure_bar": to_number(pressure_val), "status": status,
                    "new_type": False, "note": None,
                    "uploaded_by": current_user["full_name"], "uploaded_by_id": current_user["_id"],
                    "created_at": datetime.utcnow(),
                })
    if docs:
        pressure_readings_col.insert_many(docs)
    return len(docs)


def page_equipment_info(current_user):
    st.markdown("### Motor Bilgi Kartı")
    st.caption("Kaver tipi, hava filtresi, krankcase, eşanjör, dungs ve radyatör bilgileri — referans amaçlıdır.")

    if current_user["role"] in ("yonetici", "planlamaci"):
        with st.expander("📥 Excel'den güncelle"):
            st.caption("'Motor No', 'Kaver Tipi', 'Hava Filtresi', 'Krankcase', 'Eşanjör Tipi', 'Dungs', 'Radyatör Tipi', 'Not' sütunlarını içeren bir dosya yükleyin. Aynı motor için var olan bilgiyi günceller.")
            up = st.file_uploader("Excel dosyası seç", type=["xlsx"], key="eq_import_file")
            if up is not None and st.button("İçe Aktar", key="eq_import_btn"):
                try:
                    updated = import_equipment_excel(up)
                    st.success(f"{updated} motor için bilgi güncellendi.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Dosya okunamadı: {e}")

    engines = sorted(equipment_info_col.find(), key=lambda e: engine_sort_key(e["engine_name"]))
    if not engines:
        st.info("Henüz motor bilgisi eklenmemiş.")
        return

    query = st.text_input("Motor ara", placeholder="örn. AGM 12", key="eq_search")
    rows = [e for e in engines if query.lower() in e["engine_name"].lower()] if query else engines

    df = pd.DataFrame([{
        "Motor": e["engine_name"], "Kaver Tipi": e.get("kaver_tipi", ""), "Hava Filtresi": e.get("hava_filtresi", ""),
        "Krankcase": e.get("krankcase", ""), "Eşanjör Tipi": e.get("esanjor_tipi", ""),
        "Dungs": e.get("dungs", ""), "Radyatör Tipi": e.get("radyator_tipi", ""), "Not": e.get("not", ""),
    } for e in rows])
    st.dataframe(df, use_container_width=True, hide_index=True, height=560)


def import_equipment_excel(uploaded_file):
    from openpyxl import load_workbook
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    def norm(name):
        s = str(name).strip().replace("-", " ")
        return " ".join(s.split())

    headers = {}
    for c in range(1, 10):
        v = ws.cell(1, c).value
        if v:
            headers[str(v).strip().upper()] = c

    col_map = {
        "kaver_tipi": headers.get("KAVER TİPİ"), "hava_filtresi": headers.get("HAVA FİLTRESİ"),
        "krankcase": headers.get("KRANKCASE"), "esanjor_tipi": headers.get("EŞANJÖR TİPİ"),
        "dungs": headers.get("DUNGS"), "radyator_tipi": headers.get("RADYATÖR TİPİ"), "not": headers.get("NOT"),
    }
    motor_col = headers.get("MOTOR NO")
    if not motor_col:
        raise ValueError("'Motor No' sütunu bulunamadı")

    updated = 0
    for r in range(2, ws.max_row + 1):
        name_raw = ws.cell(r, motor_col).value
        if not name_raw or "AGM" not in str(name_raw).upper():
            continue
        name = norm(name_raw)
        info = {"engine_name": name}
        for key, col in col_map.items():
            if col:
                info[key] = ws.cell(r, col).value
        equipment_info_col.update_one({"_id": name}, {"$set": info}, upsert=True)
        updated += 1
    return updated


def slugify_key(label):
    """Türkçe karakterleri sadeleştirerek bir etiketten benzersiz bir anahtar üretir."""
    tr_map = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
    s = label.translate(tr_map).lower()
    s = "".join(ch if ch.isalnum() else "_" for ch in s)
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def page_maintenance_type_admin(current_user):
    if current_user["role"] != "yonetici":
        st.warning("Bu sayfa yalnızca yöneticiler içindir.")
        return
    st.markdown("### Bakım Türü Yönetimi")

    with st.expander("➕ Yeni Bakım Türü Ekle"):
        new_label = st.text_input("Bakım türü adı", key="new_type_label", placeholder="örn. Egzoz Valfi Kontrolü")
        new_period = st.number_input("Periyodik bakım saati", min_value=1.0, value=1000.0, step=100.0, key="new_type_period")
        apply_to_all = st.checkbox("Şimdi tüm motorlara uygula (motorların güncel saatini başlangıç kabul eder)",
                                    value=True, key="new_type_apply_all")
        if st.button("Bakım Türünü Ekle", key="add_type_btn"):
            if not new_label.strip():
                st.error("Lütfen bir isim girin.")
            else:
                key = slugify_key(new_label)
                if not key:
                    st.error("Geçersiz isim, lütfen farklı bir isim deneyin.")
                elif types_col.find_one({"_id": key}):
                    st.error("Bu veya çok benzer isimde bir bakım türü zaten var.")
                else:
                    engine_states = {}
                    if apply_to_all:
                        for e in engines_col.find():
                            engine_states[e["_id"]] = {"last_maintenance_hour": e["hours"], "period_hours": new_period}
                    types_col.insert_one({
                        "_id": key, "key": key, "label": new_label.strip(),
                        "default_period_hours": new_period, "engine_states": engine_states,
                    })
                    st.cache_data.clear()
                    st.success(f"'{new_label.strip()}' bakım türü eklendi.")
                    st.rerun()

    st.caption("Bir bakım türünün adını/periyodunu düzenleyebilir veya artık takip edilmesini istemediğiniz bir türü tüm geçmiş kayıtlarıyla birlikte kalıcı olarak silebilirsiniz.")

    all_types = sorted(types_col.find(), key=lambda t: t["label"])
    if not all_types:
        st.info("Tanımlı bakım türü yok.")
        return

    for t in all_types:
        record_count = records_col.count_documents({"type_key": t["key"]})
        engine_count = len(t.get("engine_states", {}))
        with st.container(border=True):
            c1, c2, c3 = st.columns([3, 1, 1])
            with c1:
                st.markdown(f"**{t['label']}**")
                st.caption(f"{engine_count} motorda tanımlı · {record_count} bakım kaydı · Varsayılan periyot: {t['default_period_hours']} sa")

            edit_key = f"edit_type_{t['_id']}"
            with c2:
                if st.button("✏️ Düzenle", key=f"edbtn_type_{t['_id']}", use_container_width=True):
                    st.session_state[edit_key] = not st.session_state.get(edit_key, False)
                    st.rerun()
            with c3:
                confirm_key = f"confirm_del_type_{t['_id']}"
                if st.session_state.get(confirm_key):
                    if st.button("⚠️ Emin misiniz?", key=f"sure_{t['_id']}", type="primary", use_container_width=True):
                        records_col.delete_many({"type_key": t["key"]})
                        types_col.delete_one({"_id": t["_id"]})
                        st.session_state[confirm_key] = False
                        st.cache_data.clear()
                        st.success(f"{t['label']} ve tüm kayıtları silindi.")
                        st.rerun()
                    if st.button("Vazgeç", key=f"cancel_del_type_{t['_id']}", use_container_width=True):
                        st.session_state[confirm_key] = False
                        st.rerun()
                else:
                    if st.button("🗑️ Sil", key=f"del_type_{t['_id']}", use_container_width=True):
                        st.session_state[confirm_key] = True
                        st.rerun()

            if st.session_state.get(edit_key):
                st.markdown("---")
                edited_label = st.text_input("Bakım türü adı", value=t["label"], key=f"ed_label_{t['_id']}")
                edited_period = st.number_input("Varsayılan periyodik bakım saati", min_value=1.0,
                                                 value=float(t["default_period_hours"]), step=100.0, key=f"ed_period_{t['_id']}")
                apply_period_all = st.checkbox("Bu periyodu, bu türü zaten kullanan tüm motorlara da uygula",
                                                value=False, key=f"ed_applyall_{t['_id']}",
                                                help="İşaretlemezseniz sadece yeni motorlar için varsayılan olarak kullanılır, mevcut motorların periyodu değişmez.")
                sc1, sc2 = st.columns(2)
                if sc1.button("Vazgeç", key=f"ed_cancel_type_{t['_id']}", use_container_width=True):
                    st.session_state[edit_key] = False
                    st.rerun()
                if sc2.button("💾 Kaydet", key=f"ed_save_type_{t['_id']}", type="primary", use_container_width=True):
                    update = {"label": edited_label.strip(), "default_period_hours": edited_period}
                    types_col.update_one({"_id": t["_id"]}, {"$set": update})
                    if apply_period_all:
                        for eng_id in t.get("engine_states", {}):
                            types_col.update_one(
                                {"_id": t["_id"]},
                                {"$set": {f"engine_states.{eng_id}.period_hours": edited_period}},
                            )
                    if edited_label.strip() != t["label"]:
                        records_col.update_many({"type_key": t["key"]}, {"$set": {"type_label": edited_label.strip()}})
                    st.session_state[edit_key] = False
                    st.cache_data.clear()
                    st.success("Bakım türü güncellendi.")
                    st.rerun()


def page_users(current_user):
    if current_user["role"] != "yonetici":
        st.warning("Bu sayfa yalnızca yöneticiler içindir.")
        return
    st.markdown("### Kullanıcılar")
    users = list(users_col.find())
    for u in users:
        with st.container(border=True):
            c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
            c1.write(f"**{u['full_name']}**")
            c2.write(u["email"])
            new_role = c3.selectbox("Rol", ROLES, index=ROLES.index(u["role"]), format_func=lambda r: ROLE_LABELS[r],
                                     key=f"role_{u['_id']}", label_visibility="collapsed")
            if new_role != u["role"]:
                users_col.update_one({"_id": u["_id"]}, {"$set": {"role": new_role}})
                st.rerun()
            active = c4.checkbox("Aktif", value=u.get("active", True), key=f"active_{u['_id']}")
            if active != u.get("active", True):
                users_col.update_one({"_id": u["_id"]}, {"$set": {"active": active}})
                st.rerun()


# ============================================================
# Ana uygulama
# ============================================================
if "user" not in st.session_state:
    st.session_state.user = None

if st.session_state.user is None:
    login_view()
else:
    user = st.session_state.user
    with st.sidebar:
        st.markdown(f"**{user['full_name']}**")
        st.caption(ROLE_LABELS.get(user["role"], user["role"]))
        if st.button("Çıkış Yap", use_container_width=True):
            st.session_state.user = None
            st.rerun()
        st.divider()

        pages = ["📊 Özet", "⏱️ Saat Güncelle", "⚙️ Motorlar", "🔧 Bakım Türleri",
                 "✅ Bakım Tamamla", "🗓️ Bakım Tarihi Tahmini", "🧪 Yağ Analizleri", "📉 Karter Fark Basıncı",
                 "📋 Motor Bilgi Kartı", "📜 Bakım Kayıtları",
                 "📈 Saat Geçmişi", "📊 Bakım Aralıkları", "📥 Excel"]
        if user["role"] == "yonetici":
            pages.append("👥 Kullanıcılar")
            pages.append("🗑️ Bakım Türü Yönetimi")
        choice = st.radio("Menü", pages, label_visibility="collapsed")

    if choice == "📊 Özet":
        page_dashboard()
    elif choice == "⏱️ Saat Güncelle":
        if user["role"] in ("yonetici", "planlamaci"):
            page_hours_update()
        else:
            st.warning("Bu işlem için yönetici veya planlamacı yetkisi gerekir.")
    elif choice == "⚙️ Motorlar":
        page_engines()
    elif choice == "🔧 Bakım Türleri":
        page_types()
    elif choice == "✅ Bakım Tamamla":
        page_complete_maintenance(user)
    elif choice == "🗓️ Bakım Tarihi Tahmini":
        page_maintenance_forecast()
    elif choice == "🧪 Yağ Analizleri":
        page_oil_analyses(user)
    elif choice == "📉 Karter Fark Basıncı":
        page_pressure_readings(user)
    elif choice == "📋 Motor Bilgi Kartı":
        page_equipment_info(user)
    elif choice == "📜 Bakım Kayıtları":
        page_records(user)
    elif choice == "📈 Saat Geçmişi":
        page_hours_history()
    elif choice == "📊 Bakım Aralıkları":
        page_intervals()
    elif choice == "📥 Excel":
        page_export()
    elif choice == "👥 Kullanıcılar":
        page_users(user)
    elif choice == "🗑️ Bakım Türü Yönetimi":
        page_maintenance_type_admin(user)
